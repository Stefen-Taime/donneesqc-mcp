"""Telecharge une ressource non-DataStore (CSV, JSON, GeoJSON, XLSX)."""

import csv
import io
import json
import logging

from helpers import ckan_api
from helpers.config import DQ_API_URL
from helpers.http_client import fetch_bytes
from tools import mcp

logger = logging.getLogger(__name__)

# Formats supportes pour le parsing
_TEXT_FORMATS = {"csv", "json", "geojson", "txt", "xml"}
_BINARY_FORMATS = {"xlsx", "xls"}
_ALL_FORMATS = _TEXT_FORMATS | _BINARY_FORMATS


@mcp.tool()
async def download_resource(
    resource_id: str,
    max_rows: int = 100,
) -> str:
    """Telecharge et lit une ressource non-DataStore de Donnees Quebec.

    Utile quand datastore_active=false. Supporte CSV, JSON, GeoJSON.
    Pour les fichiers XLSX, seuls les headers sont retournes (necessite openpyxl).

    Parameters:
        resource_id: ID de la ressource a telecharger.
        max_rows: Nombre maximal de lignes a retourner (defaut 100, max 500).
    """
    try:
        max_rows = max(1, min(max_rows, 500))

        # Recuperer les metadonnees de la ressource
        res = await ckan_api.resource_show(DQ_API_URL, resource_id)
        url = res.get("url")
        fmt = (res.get("format") or "").lower().strip()

        if not url:
            return json.dumps({"error": "Aucune URL de telechargement pour cette ressource."}, ensure_ascii=False)

        if fmt not in _ALL_FORMATS:
            return json.dumps(
                {
                    "error": f"Format '{fmt}' non supporte. Formats acceptes: {', '.join(sorted(_ALL_FORMATS))}.",
                    "url": url,
                    "tip": "Vous pouvez telecharger le fichier directement via l'URL.",
                },
                ensure_ascii=False,
            )

        raw = await fetch_bytes(url)

        if fmt == "csv":
            return _parse_csv(raw, max_rows, res)
        if fmt in ("json", "geojson"):
            return _parse_json(raw, max_rows, res)
        if fmt in ("xlsx", "xls"):
            return _parse_xlsx_headers(raw, res)
        text = raw.decode("utf-8", errors="replace")[:5000]
        return json.dumps(
            {
                    "resource_id": resource_id,
                    "format": fmt,
                    "preview": text,
                    "note": "Apercu brut (5000 premiers caracteres)",
                },
            ensure_ascii=False,
            indent=2,
        )

    except Exception:
        logger.exception("Erreur dans download_resource")
        return json.dumps(
            {"error": f"Impossible de telecharger la ressource '{resource_id}'."},
            ensure_ascii=False,
        )


def _parse_csv(raw: bytes, max_rows: int, res: dict) -> str:
    """Parse un fichier CSV et retourne les premieres lignes."""
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for i, row in enumerate(reader):
        if i >= max_rows:
            break
        rows.append(dict(row))

    return json.dumps(
        {
            "resource_id": res.get("id"),
            "name": res.get("name"),
            "format": "CSV",
            "total_returned": len(rows),
            "fields": list(rows[0].keys()) if rows else [],
            "records": rows,
            "truncated": len(rows) == max_rows,
        },
        ensure_ascii=False,
        indent=2,
    )


def _parse_json(raw: bytes, max_rows: int, res: dict) -> str:
    """Parse un fichier JSON/GeoJSON."""
    data = json.loads(raw)

    # GeoJSON
    if isinstance(data, dict) and "features" in data:
        features = data["features"][:max_rows]
        return json.dumps(
            {
                "resource_id": res.get("id"),
                "name": res.get("name"),
                "format": "GeoJSON",
                "total_features": len(data["features"]),
                "returned": len(features),
                "features": features,
                "truncated": len(data["features"]) > max_rows,
            },
            ensure_ascii=False,
            indent=2,
        )

    # JSON array
    if isinstance(data, list):
        return json.dumps(
            {
                "resource_id": res.get("id"),
                "name": res.get("name"),
                "format": "JSON",
                "total_records": len(data),
                "returned": min(len(data), max_rows),
                "records": data[:max_rows],
                "truncated": len(data) > max_rows,
            },
            ensure_ascii=False,
            indent=2,
        )

    # JSON object
    return json.dumps(
        {
            "resource_id": res.get("id"),
            "name": res.get("name"),
            "format": "JSON",
            "data": data,
        },
        ensure_ascii=False,
        indent=2,
    )


def _parse_xlsx_headers(raw: bytes, res: dict) -> str:
    """Retourne les headers d'un fichier XLSX (sans openpyxl en dependance obligatoire)."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheets_info = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(max_row=6, values_only=True))
            headers = [str(c) if c is not None else "" for c in rows[0]] if rows else []
            sample = [
                [str(c) if c is not None else "" for c in row]
                for row in rows[1:]
            ]
            sheets_info.append({"sheet": name, "headers": headers, "sample_rows": sample})
        wb.close()

        return json.dumps(
            {
                "resource_id": res.get("id"),
                "name": res.get("name"),
                "format": "XLSX",
                "sheets": sheets_info,
                "note": "Apercu headers + 5 lignes. Pour lecture complete, telecharger via l'URL.",
            },
            ensure_ascii=False,
            indent=2,
        )
    except ImportError:
        return json.dumps(
            {
                "resource_id": res.get("id"),
                "format": "XLSX",
                "error": "openpyxl non installe. Installez-le avec: pip install openpyxl",
                "url": res.get("url"),
            },
            ensure_ascii=False,
        )
